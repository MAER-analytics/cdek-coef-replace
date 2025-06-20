import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from io import BytesIO
import os

def process_file():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not filepath:
        return

    try:
        wb = load_workbook(filepath, data_only=False, keep_links=False)
        sheet_name = next((name for name in wb.sheetnames if "СДЭК" in name.upper()), None)

        if not sheet_name:
            messagebox.showerror("Ошибка", "Не найден лист, содержащий 'СДЭК' в названии.")
            return

        ws = wb[sheet_name]

        # Ищем ячейку с заголовком
        target_cell = None
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=30):
            for cell in row:
                if str(cell.value).strip().lower() == "локальный коэффициент":
                    target_cell = cell
                    break
            if target_cell:
                break

        if not target_cell:
            messagebox.showerror("Ошибка", "Заголовок 'Локальный коэффициент' не найден.")
            return

        col_letter = target_cell.column_letter
        row = target_cell.row + 1

        # Пропускаем пустые строки после заголовка
        while ws[f"{col_letter}{row}"].value in [None, ""]:
            row += 1

        changed_count = 0

        # Замена значений
        while True:
            cell = ws[f"{col_letter}{row}"]
            raw_value = cell.value
            if raw_value is None or str(raw_value).strip() == "":
                break

            value_str = str(raw_value).replace(" ", "").replace(",", ".")
            try:
                value_float = float(value_str)
            except ValueError:
                row += 1
                continue

            if value_float == 1.5:
                cell.value = "1,2"
                changed_count += 1
            elif value_float == 2.0:
                cell.value = "1,4"
                changed_count += 1
            elif value_float == 1.0:
                cell.value = "1"
                changed_count += 1

            row += 1

        # Сохранение
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            wb.save(save_path)
            messagebox.showinfo("Готово", f"Файл сохранён. Заменено значений: {changed_count}")

    except Exception as e:
        messagebox.showerror("Ошибка", str(e))


# UI
root = tk.Tk()
root.title("CDEK Коэффициенты")

button = tk.Button(root, text="Выбрать и обработать файл", command=process_file, width=30, height=2)
button.pack(pady=20)

root.mainloop()
