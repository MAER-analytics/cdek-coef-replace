import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def select_file():
    filepath = filedialog.askopenfilename(
        title="Выберите Excel-файл",
        filetypes=[("Excel-файлы", "*.xlsx")]
    )
    if filepath:
        entry_var.set(filepath)

def process_file():
    filepath = entry_var.get()
    if not filepath:
        messagebox.showwarning("Файл не выбран", "Пожалуйста, выберите Excel-файл.")
        return

    try:
        wb = load_workbook(filepath, data_only=False, keep_links=False)
        sheet_name = next((name for name in wb.sheetnames if "СДЭК" in name.upper()), None)

        if not sheet_name:
            messagebox.showerror("Ошибка", "Не найден лист, содержащий 'СДЭК' в названии.")
            return

        ws = wb[sheet_name]

        target_cell = None
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=30):
            for cell in row:
                if str(cell.value).strip().lower() == "локальный коэффициент":
                    target_cell = cell
                    break
            if target_cell:
                break

        if not target_cell:
            messagebox.showerror("Ошибка", "Не найден заголовок 'Локальный коэффициент'.")
            return

        col_letter = target_cell.column_letter
        row = target_cell.row + 1

        while ws[f"{col_letter}{row}"].value in [None, ""]:
            row += 1

        changed_count = 0

        while True:
            cell = ws[f"{col_letter}{row}"]
            raw_value = cell.value

            if raw_value is None or str(raw_value).strip() == "":
                break

            value_str = str(raw_value).replace(" ", "").replace(",", ".")
            try:
                value_float = float(value_str)
            except ValueError:
                row += 1
                continue

            if value_float == 1.5:
                cell.value = "1,2"
                changed_count += 1
            elif value_float == 2.0:
                cell.value = "1,4"
                changed_count += 1
            elif value_float == 1.0:
                cell.value = "1"
                changed_count += 1

            row += 1

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel-файлы", "*.xlsx")],
            title="Сохранить как..."
        )

        if save_path:
            wb.save(save_path)
            messagebox.showinfo("Готово", f"Файл успешно сохранён.\nЗаменено значений: {changed_count}")
        else:
            messagebox.showinfo("Отмена", "Сохранение отменено.")

    except Exception as e:
        messagebox.showerror("Ошибка обработки", str(e))


# --- Интерфейс ---
root = tk.Tk()
root.title("Автозамена коэффициентов (СДЭК)")
root.geometry("460x240")
root.resizable(False, False)

entry_var = tk.StringVar()

title_label = tk.Label(root, text="Excel: автозамена коэффициентов", font=("Helvetica", 14, "bold"))
title_label.pack(pady=(20, 5))

desc_label = tk.Label(
    root,
    text="Выберите Excel-файл с листом 'СДЭК'.\nПрограмма заменит 1.5 → 1.2 и 2 → 1.4 в нужном столбце.",
    font=("Helvetica", 10),
    justify="center"
)
desc_label.pack(pady=5)

frame = tk.Frame(root)
frame.pack(pady=10)

entry = tk.Entry(frame, textvariable=entry_var, width=40)
entry.pack(side=tk.LEFT, padx=(0, 10))

select_button = tk.Button(frame, text="📁 Выбрать файл", command=select_file)
select_button.pack(side=tk.LEFT)

process_button = tk.Button(root, text="🔁 Обработать и сохранить", command=process_file, width=30, height=2)
process_button.pack(pady=15)

root.mainloop()
