import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import os

def process_file():
    filepath = filedialog.askopenfilename(
        title="Выберите Excel-файл",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not filepath:
        return

    try:
        wb = load_workbook(filepath, data_only=False, keep_links=False)
        sheet_name = next((name for name in wb.sheetnames if "СДЭК" in name.upper()), None)

        if not sheet_name:
            messagebox.showerror("Ошибка", "Не найден лист с 'СДЭК' в названии.")
            return

        ws = wb[sheet_name]

        # Поиск заголовка
        target_cell = None
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=30):
            for cell in row:
                if str(cell.value).strip().lower() == "локальный коэффициент":
                    target_cell = cell
                    break
            if target_cell:
                break

        if not target_cell:
            messagebox.showerror("Ошибка", "Не найден заголовок 'Локальный коэффициент'.")
            return

        col_letter = target_cell.column_letter
        row = target_cell.row + 1

        while ws[f"{col_letter}{row}"].value in [None, ""]:
            row += 1

        changed_count = 0

        while True:
            cell = ws[f"{col_letter}{row}"]
            raw_value = cell.value
            if raw_value is None or str(raw_value).strip() == "":
                break

            value_str = str(raw_value).replace(" ", "").replace(",", ".")
            try:
                value_float = float(value_str)
            except ValueError:
                row += 1
                continue

            if value_float == 1.5:
                cell.value = "1,2"
                changed_count += 1
            elif value_float == 2.0:
                cell.value = "1,4"
                changed_count += 1
            elif value_float == 1.0:
                cell.value = "1"
                changed_count += 1

            row += 1

        # Предложить путь для сохранения
        base, ext = os.path.splitext(filepath)
        default_path = base + "_обновлено.xlsx"

        save_path = filedialog.asksaveasfilename(
            title="Сохраните обновлённый файл",
            defaultextension=".xlsx",
            initialfile=os.path.basename(default_path),
            filetypes=[("Excel files", "*.xlsx")]
        )

        if save_path:
            wb.save(save_path)
            messagebox.showinfo("Готово", f"Файл сохранён!\nЗаменено значений: {changed_count}")

    except Exception as e:
        messagebox.showerror("Ошибка", str(e))


# Интерфейс
root = tk.Tk()
root.title("🛠 Автозамена коэффициентов в Excel (СДЭК)")
root.geometry("450x250")
root.resizable(False, False)

title = tk.Label(
    root, 
    text="Добро пожаловать!",
    font=("Arial", 16, "bold"),
    pady=10
)
title.pack()

description = tk.Label(
    root,
    text="Это приложение автоматически заменяет коэффициенты\n1,5 → 1,2 и 2 → 1,4 в Excel-файлах от СДЭК.",
    font=("Arial", 11),
    justify="center",
    wraplength=400
)
description.pack(pady=5)

button = tk.Button(
    root,
    text="📁 Выбрать Excel-файл",
    font=("Arial", 12),
    command=process_file,
    bg="#4CAF50",
    fg="white",
    padx=20,
    pady=10
)
button.pack(pady=25)

footer = tk.Label(
    root,
    text="Автор: Maksim\nПоддерживаются только .xlsx файлы",
    font=("Arial", 9),
    fg="gray"
)
footer.pack(side="bottom", pady=5)

root.mainloop()
