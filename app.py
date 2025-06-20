import streamlit as st
import openpyxl
from io import BytesIO

st.title("Автозамена коэффициентов в Excel (СДЭК)")

uploaded_file = st.file_uploader("Загрузите Excel-файл", type=["xlsx"])

if uploaded_file:
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=False, keep_links=False)
        # Пытаемся найти лист с именем, содержащим "СДЭК"
        sheet_name = next((name for name in wb.sheetnames if "СДЭК" in name.upper()), None)

        if not sheet_name:
            st.error("Не найден лист, содержащий 'СДЭК' в названии.")
        else:
            ws = wb[sheet_name]

            # Ищем ячейку с заголовком "Локальный коэффициент"
            target_cell = None
            for row in ws.iter_rows(min_row=1, max_row=100, max_col=30):
                for cell in row:
                    if str(cell.value).strip().lower() == "локальный коэффициент":
                        target_cell = cell
                        break
                if target_cell:
                    break

            if not target_cell:
                st.error("Заголовок 'Локальный коэффициент' не найден.")
            else:
                col_letter = target_cell.column_letter
                row = target_cell.row + 1

                # Пропускаем пустые строки после заголовка
                while ws[f"{col_letter}{row}"].value in [None, ""]:
                    row += 1

                changed_count = 0

                # Заменяем значения до первого пустого значения
                while True:
                    cell = ws[f"{col_letter}{row}"]
                    raw_value = cell.value

                    if raw_value is None or str(raw_value).strip() == "":
                        break

                    # Приведение к числу (русская локаль поддержана через замены)
                    value_str = str(raw_value).replace(" ", "").replace(",", ".")
                    try:
                        value_float = float(value_str)
                    except ValueError:
                        row += 1
                        continue

                    if value_float == 1.5:
                        cell.value = "1,2"
                        changed_count += 1
                    elif value_float == 2.0:
                        cell.value = "1,4"
                        changed_count += 1
                    elif value_float == 1.0:
                        cell.value = "1"
                        changed_count += 1
                    # Если другое значение — ничего не делаем

                    row += 1

                # Сохраняем файл
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                st.success(f"Успешно заменено значений: {changed_count}")
                st.download_button(
                    label="📥 Скачать обновлённый файл",
                    data=output,
                    file_name="updated_sdek.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"Ошибка: {e}")
